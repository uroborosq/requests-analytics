import openpyxl
from openpyxl import styles


class DaySchedule(object):
    def __init__(self, data, date):
        self.date = date
        wb = openpyxl.Workbook()
        worksheet = wb['Sheet']
        worksheet['B2'] = 'Дата'
        worksheet['C2'] = str(date.strftime('%d-%m-%Y'))
        for i in ['A', 'B', 'C', 'D', 'E']:
            worksheet.merge_cells(i + '4:' + i + '5')
        worksheet.merge_cells('F4:K4')
        worksheet.merge_cells('L4:M4')

        tmp = {'A4': '№ п/п', 'B4': '№ наряд-заказа', 'C4': 'Исполнитель', 'D4': 'Гарантия/внутр/платно',
               'E4': 'Сколько ЗИП потратил', 'F4': 'В электронном виде прислал', 'F5': 'Описание работ',
               'G5': 'Наряд-заказ',
               'H5': 'Гарталон', 'I5': 'Сер номер', 'J5': 'Деталь', 'K5': 'Шильда', 'L4': 'Сдал в бумажном виде',
               'L5': 'Наряд-заказ', 'M5': 'Накладные'}
        for i in tmp:
            worksheet[i] = tmp[i]

        tmp = {'A': 4, 'B': 11, 'C': 30, 'D': 10, 'E': 12,
               'F': 35, 'G': 6, 'H': 6, 'I': 7, 'J': 35, 'L': 6, 'M': 6, 'K': 6}

        worksheet['F4'].alignment = styles.Alignment(horizontal="center")
        worksheet['L4'].alignment = styles.Alignment(horizontal="center")
        for i in tmp:
            cl = worksheet.column_dimensions[i]
            cl.width = tmp[i]

        counter = 1
        for key, row in data.iterrows():
            worksheet['A' + str(counter + 5)] = counter  # номер
            worksheet['B' + str(counter + 5)] = key  # наряд заказ
            worksheet['D' + str(counter + 5)] = row['Ссылка.Гарантийность']
            tmp = ''
            for k in row['Ссылка.Исполнители.Исполнитель']:
                if k is not None:
                    tmp += k + ' \n '
            if row['Ссылка.Адрес'] == '':
                tmp += 'Адрес не указан'
            else:
                tmp += row['Ссылка.Адрес']
            worksheet['C' + str(counter + 5)] = tmp
            counter += 1
        worksheet.row_dimensions[4].height = 40
        worksheet.row_dimensions[5].height = 40
        for i in range(counter + 2):
            rw = worksheet.row_dimensions[i + 6]
            rw.height = 60

        tmp = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
        for i in range(13):
            for j in range(counter + 2):
                tmp.append(tmp[i] + str(j + 6))

        for i in ['B2', 'C2', 'A4', 'B4', 'C4', 'D4', 'E4', 'F4', 'F5', 'G5', 'H5', 'I5', 'J5', 'K5', 'L4', 'L5', 'M5',
                  'A5', 'B5', 'C5', 'D5', 'E5', 'M4', 'G4', 'H4', 'I4', 'J4', 'K4'] + tmp[13:]:
            worksheet[i].border = styles.Border(top=styles.Side(border_style="thin", color="000000"),
                                                left=styles.Side(
                                                    border_style="thin", color="000000"),
                                                right=styles.Side(
                                                    border_style="thin", color="000000"),
                                                bottom=styles.Side(border_style="thin", color="000000"))
        for i in tmp[13:]:
            worksheet[i].alignment = styles.Alignment(
                horizontal='center', vertical='center')

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = cell.alignment.copy(wrapText=True)

        wb.save('DaySchedule' + str(date.strftime('%d-%m-%Y')) + '.xlsx')

    def get(self):
        return 'DaySchedule' + str(self.date.strftime('%d-%m-%Y')) + '.xlsx'